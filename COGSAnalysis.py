import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils.dataframe import dataframe_to_rows

class COGSAnalysis:
    def __init__(self):
        self.BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        self.INPUT_DIR = os.path.join(self.BASE_DIR, 'COGS Inputs')
        self.OUTPUT_DIR = os.path.join(self.BASE_DIR, 'COGS Outputs')
        self.PUBLIC_DIR = r"C:\City Experience\Public Data Base"
        self.cost_columns = ['Coordinator', 'Food', 'Guide', 'Headsets', 'Misc', 'Tickets', 'Transport']
        
        # Ensure output directory exists
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)
    
    def get_season(self, date):
        """Determine season based on quarter"""
        month = pd.to_datetime(date).month
        return 'High Season' if month in [4,5,6,7,8,9] else 'Low Season'
    
    def add_per_pax_columns(self, df):
        """Add per pax columns for cost metrics"""
        for col in self.cost_columns + ['COGS']:
            df[f'{col} per Pax'] = df[col] / df['Actual Pax'].replace(0, np.nan)
        return df
    
    def prepare_vendor_data(self):
        """Prepare vendor data from AP Prediction.xlsx"""
        df = pd.read_excel(os.path.join(self.INPUT_DIR, 'AP Prediction.xlsx'), sheet_name='Cost Export')
        df = df[['Stage ID', 'Vendor', 'Event Date', 'Event', 'Event ID'] + self.cost_columns]
        
        # Calculate COGS
        df['COGS'] = df[self.cost_columns].sum(axis=1)
        
        # Add Season and Year
        df['Season'] = df['Event Date'].apply(self.get_season)
        df['Year'] = pd.to_datetime(df['Event Date']).dt.year
        
        # Filter for 2024 and later
        df = df[df['Year'] >= 2024]
        
        # Fill missing Vendor values
        df['Vendor'] = df['Vendor'].fillna('Blank Vendor')
        
        return df
    
    def prepare_event_accounting_data(self):
        """Prepare event accounting data from Revenue Basic Data.xlsx"""
        df = pd.read_excel(os.path.join(self.INPUT_DIR, 'Revenue Basic Data.xlsx'), sheet_name='Event Accounting')
        return df[['Stage ID', 'Actual Pax']]
    
    def prepare_merged_data(self):
        """Prepare merged vendor and stage ID level data"""
        vendor_data = self.prepare_vendor_data()
        event_accounting = self.prepare_event_accounting_data()
        
        # Vendor level merged data
        vendor_merged = vendor_data.merge(event_accounting, on='Stage ID', how='left')
        vendor_merged.to_excel(os.path.join(self.OUTPUT_DIR, 'vendor_level_data.xlsx'), index=False)
        
        # Stage ID level data
        group_cols = ['Stage ID', 'Event Date', 'Event', 'Event ID', 'Season', 'Year']
        stage_id_data = vendor_data.groupby(group_cols)[self.cost_columns + ['COGS']].sum().reset_index()
        stage_id_data = stage_id_data.merge(event_accounting, on='Stage ID', how='left')
        stage_id_data.to_excel(os.path.join(self.OUTPUT_DIR, 'stage_id_level_data.xlsx'), index=False)
        
        return vendor_merged, stage_id_data
    
    def format_worksheet(self, worksheet, df, highlight_column=None, highlight_type='std'):
        """Format Excel worksheet with conditional formatting using openpyxl"""
        # Write headers
        for col_num, col_name in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_num).value = col_name
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                worksheet.cell(row=r_idx, column=c_idx).value = value
        
        if highlight_column:
            col_idx = df.columns.get_loc(highlight_column) + 1
            values = [worksheet.cell(row=r, column=col_idx).value for r in range(2, len(df) + 2)]
            values = [v for v in values if pd.notnull(v)]
            
            if highlight_type == 'std':
                # Highlight top 10 standard deviation
                std_threshold = sorted(values, reverse=True)[:10]
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                red_font = Font(color='9C0006')
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value in std_threshold:
                        cell.fill = red_fill
                        cell.font = red_font
            elif highlight_type == 'change':
                # Highlight top 10 percentage change
                change_threshold = sorted(values, reverse=True)[:10]
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                red_font = Font(color='9C0006')
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value in change_threshold:
                        cell.fill = red_fill
                        cell.font = red_font
            elif highlight_type == 'minmax':
                # Highlight min and max values
                max_val = max(values)
                min_val = min(values)
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red_font = Font(color='9C0006')
                green_font = Font(color='006100')
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value == max_val:
                        cell.fill = red_fill
                        cell.font = red_font
                    elif cell.value == min_val:
                        cell.fill = green_fill
                        cell.font = green_font
    
    def per_pax_cogs_analysis(self, stage_id_data):
        """Perform Per Pax COGS analysis"""
        # Aggregate data
        group_cols = ['Event', 'Event ID', 'Season', 'Year']
        agg_data = stage_id_data.groupby(group_cols)[self.cost_columns + ['COGS', 'Actual Pax']].sum().reset_index()
        
        # Add per pax columns
        agg_data = self.add_per_pax_columns(agg_data)
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Longitudinal analysis
        for col in self.cost_columns + ['COGS']:
            pivot = agg_data.pivot_table(
                values=f'{col} per Pax',
                index=['Event', 'Event ID', 'Season'],
                columns='Year'
            ).reset_index()
            
            pivot['Change %'] = ((pivot[2025] - pivot[2024]) / pivot[2024] * 100).replace([np.inf, -np.inf], np.nan)
            pivot = pivot.sort_values('Change %', ascending=False)
            
            ws = wb.create_sheet(f'{col} Longitudinal')
            self.format_worksheet(ws, pivot, 'Change %', 'change')
        
        # Statistical analysis
        stats_cols = [f'{col} per Pax' for col in self.cost_columns + ['COGS']]
        stats = agg_data.groupby(['Event', 'Event ID', 'Season'])[stats_cols].agg(['mean', 'std', 'max', 'min']).reset_index()
        
        for col in stats_cols:
            stat_df = stats[[('Event', ''), ('Event ID', ''), ('Season', ''), (col, 'mean'), (col, 'std'), (col, 'max'), (col, 'min')]]
            stat_df.columns = ['Event', 'Event ID', 'Season', 'Mean', 'Std', 'Max', 'Min']
            ws = wb.create_sheet(f'{col} Stats')
            self.format_worksheet(ws, stat_df, 'Std', 'std')
        
        wb.save(os.path.join(self.OUTPUT_DIR, 'per_pax_cogs_analysis.xlsx'))
    
    def per_vendor_cogs_analysis(self, vendor_data):
        """Perform Per Vendor COGS analysis"""
        # Add per pax columns
        vendor_data = self.add_per_pax_columns(vendor_data)
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Longitudinal analysis
        for col in self.cost_columns + ['COGS']:
            pivot = vendor_data.pivot_table(
                values=f'{col} per Pax',
                index=['Event', 'Event ID', 'Season'],
                columns='Year'
            ).reset_index()
            
            pivot['Change %'] = ((pivot[2025] - pivot[2024]) / pivot[2024] * 100).replace([np.inf, -np.inf], np.nan)
            pivot = pivot.sort_values('Change %', ascending=False)
            
            ws = wb.create_sheet(f'{col} Longitudinal')
            self.format_worksheet(ws, pivot, 'Change %', 'change')
        
        # Cross-sectional analysis
        cross_data = vendor_data.groupby(['Vendor', 'Season', 'Year'])[self.cost_columns + ['COGS', 'Actual Pax']].sum().reset_index()
        cross_data = self.add_per_pax_columns(cross_data)
        
        for col in self.cost_columns + ['COGS']:
            pivot = cross_data.pivot_table(
                values=f'{col} per Pax',
                index=['Vendor'],
                columns=['Year', 'Season']
            ).reset_index()
            
            ws = wb.create_sheet(f'{col} Cross-sectional')
            self.format_worksheet(ws, pivot, f'{col} per Pax', 'minmax')
        
        # Statistical analysis - Longitudinal
        stats_cols = [f'{col} per Pax' for col in self.cost_columns + ['COGS']]
        long_stats = vendor_data.groupby(['Event', 'Event ID', 'Season'])[stats_cols].agg(['mean', 'std', 'max', 'min']).reset_index()
        
        for col in stats_cols:
            stat_df = long_stats[[('Event', ''), ('Event ID', ''), ('Season', ''), (col, 'mean'), (col, 'std'), (col, 'max'), (col, 'min')]]
            stat_df.columns = ['Event', 'Event ID', 'Season', 'Mean', 'Std', 'Max', 'Min']
            ws = wb.create_sheet(f'{col} Long Stats')
            self.format_worksheet(ws, stat_df, 'Std', 'std')
        
        # Statistical analysis - Cross-sectional
        cross_stats = cross_data.groupby(['Year', 'Season'])[stats_cols].agg(['mean', 'std', 'max', 'min']).reset_index()
        
        for col in stats_cols:
            stat_df = cross_stats[[('Year', ''), ('Season', ''), (col, 'mean'), (col, 'std'), (col, 'max'), (col, 'min')]]
            stat_df.columns = ['Year', 'Season', 'Mean', 'Std', 'Max', 'Min']
            ws = wb.create_sheet(f'{col} Cross Stats')
            self.format_worksheet(ws, stat_df, 'Std', 'std')
        
        wb.save(os.path.join(self.OUTPUT_DIR, 'per_vendor_cogs_analysis.xlsx'))
    
    def run_analysis(self):
        """Run complete COGS analysis"""
        vendor_data, stage_id_data = self.prepare_merged_data()
        self.per_pax_cogs_analysis(stage_id_data)
        self.per_vendor_cogs_analysis(vendor_data)

if __name__ == "__main__":
    analysis = COGSAnalysis()
    analysis.run_analysis()
